# Copyright (c) 2025 <Tretyakov Vyacheslav>
# See README.md file in the project root for full license information.

import pytz
import json
import atexit
import threading
import requests
import webbrowser
import urllib3
import logging
from logging.handlers import RotatingFileHandler
import os
import sys
from flask import Flask, request, render_template, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, text
from sqlalchemy.exc import IntegrityError
from openpyxl.styles import Font
from openpyxl import Workbook
from datetime import datetime, timezone, timedelta
import time

# Отладка пути поиска модулей
print("sys.path:", sys.path)

# Попытка импорта модулей
try:
    from netbox import fetch_storage_from_netbox
    from huawei import perform_huawei_discovery
    from netapp import perform_netapp_discovery  
    from ibm import perform_ibm_discovery
except ModuleNotFoundError as e:
    print(f"Ошибка импорта модулей: {e}")
    print("Убедитесь, что модули находятся в правильной директории.")
    raise

LOCKFILE = 'browser_opened.lock'

# Отключаем предупреждения для самоподписанных сертификатов
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# -------------------------------
# Настройка логирования
# -------------------------------
LOG_FILE = "storage_discovery.log"
MAX_LOG_SIZE = 5 * 1024 * 1024

# Проверяем размер лог-файла
if os.path.exists(LOG_FILE):
    try:
        file_size = os.path.getsize(LOG_FILE)
        if file_size > MAX_LOG_SIZE:
            print(f"Файл {LOG_FILE} превысил размер {MAX_LOG_SIZE} байт "
                  f"(было {file_size} байт). Удаление файла.")
            os.remove(LOG_FILE)
    except Exception as e:
        print(f"Ошибка при проверке размера файла {LOG_FILE}: {str(e)}")

# Настройка логгера
logger = logging.getLogger("StorageDiscovery")
logger.setLevel(logging.INFO)

# Убираем старые обработчики если они есть
for handler in logger.handlers[:]:
    logger.removeHandler(handler)

# Создаем новый обработчик
formatter = logging.Formatter(
    "%(name)s - %(asctime)s - %(funcName)s - %(levelname)s - %(message)s"
)

try:
    file_handler = RotatingFileHandler(LOG_FILE, maxBytes=MAX_LOG_SIZE, backupCount=0)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
except Exception as e:
    print(f"Не удалось создать файловый логгер: {e}")
    # Добавляем консольный логгер как запасной вариант
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

# Отключаем логирование Flask для чистоты вывода
logging.getLogger('werkzeug').setLevel(logging.WARNING)

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///storage.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# URL NetBox (можно сделать настраиваемым через конфиг)
NETBOX_URL = "https://netbox.mng.sbercloud.tech/"
# Константа часового пояса
MOSCOW_TZ = pytz.timezone('Europe/Moscow')


# -------------------------------
# Вспомогательные функции для работы со временем
# -------------------------------
def get_moscow_time():
    """Возвращает текущее время в московском часовом поясе как naive datetime"""
    utc_now = datetime.utcnow()
    utc_dt = pytz.utc.localize(utc_now)
    moscow_dt = utc_dt.astimezone(MOSCOW_TZ)
    return moscow_dt.replace(tzinfo=None)


def utc_to_moscow(utc_datetime):
    """Конвертирует UTC время в московское"""
    if utc_datetime is None:
        return None
    if isinstance(utc_datetime, str):
        try:
            if 'T' in utc_datetime:
                utc_datetime = datetime.fromisoformat(utc_datetime.replace('Z', ''))
            else:
                utc_datetime = datetime.strptime(utc_datetime, '%Y-%m-%d %H:%M:%S')
        except:
            return utc_datetime
    
    if utc_datetime.tzinfo is not None:
        moscow_dt = utc_datetime.astimezone(MOSCOW_TZ)
        return moscow_dt.replace(tzinfo=None)
    else:
        utc_dt = pytz.utc.localize(utc_datetime)
        moscow_dt = utc_dt.astimezone(MOSCOW_TZ)
        return moscow_dt.replace(tzinfo=None)


def datetime_serializer(obj):
    """Кастомный сериализатор для datetime объектов"""
    if isinstance(obj, datetime):
        return obj.strftime('%Y-%m-%d %H:%M:%S')
    raise TypeError(f"Object of type {type(obj)} is not JSON serializable")


# Функции времени для использования в шаблонах
@app.template_filter('moscow_time')
def moscow_time_filter(dt):
    """Фильтр для отображения времени в московском часовом поясе"""
    if dt is None:
        return 'Не обновлялось'
    
    if isinstance(dt, str):
        try:
            if 'T' in dt:
                dt = datetime.fromisoformat(dt.replace('Z', ''))
            else:
                dt = datetime.strptime(dt, '%Y-%m-%d %H:%M:%S')
        except:
            return dt
    
    if isinstance(dt, datetime):
        return dt.strftime('%d.%m.%Y %H:%M')
    
    return str(dt)


@app.context_processor
def utility_processor():
    return dict(utc_to_moscow=utc_to_moscow)


# -------------------------------
# Класс: Список систем хранения
# -------------------------------
class StorageSystem(db.Model):
    __tablename__ = "storage_system"
    id = db.Column(db.Integer, primary_key=True)
    table_name = db.Column(db.String(10), nullable=False)
    storage_name = db.Column(db.String(50), nullable=False)
    ip_address = db.Column(db.String(50), nullable=False)
    storage_role = db.Column(db.String(50), nullable=True)
    storage_status = db.Column(db.String(50), nullable=True)
    vendor = db.Column(db.String(20), nullable=True, default='Huawei')
    last_updated = db.Column(db.DateTime, nullable=True)
    __table_args__ = (
        db.UniqueConstraint('table_name', 'storage_name', 'ip_address', 
                          name='unique_storage_system'),
    )

    def __repr__(self):
        return (f"<StorageSystem {self.vendor} {self.table_name} "
                f"{self.storage_name} {self.ip_address}>")


# -------------------------------
# Класс: Результаты Discovery
# -------------------------------
class DiscoveryResult(db.Model):
    __tablename__ = "discovery_result"
    id = db.Column(db.Integer, primary_key=True)
    table_name = db.Column(db.String(10), nullable=False)
    data_json = db.Column(db.Text, nullable=False)
    timestamp = db.Column(db.DateTime, default=get_moscow_time)

    def __repr__(self):
        return f"<DiscoveryResult {self.table_name} {self.timestamp}>"


# Инициализация базы данных
with app.app_context():
    db.create_all()
    
    # Проверяем существование колонки vendor более безопасным способом
    try:
        # Пытаемся выполнить запрос, который использует колонку vendor
        db.session.execute(text('SELECT vendor FROM storage_system LIMIT 1')).fetchone()
        print("Колонка vendor уже существует")
    except Exception:
        try:
            # Если колонка не существует, добавляем её
            db.session.execute(text('ALTER TABLE storage_system ADD COLUMN vendor VARCHAR(20) DEFAULT "Huawei"'))
            db.session.commit()
            print("Добавлена колонка vendor")
        except Exception as e:
            print(f"Ошибка при добавлении колонки vendor: {e}")
            db.session.rollback()


# -------------------------------
# Удаление дубликатов при запуске
# -------------------------------
def remove_all_duplicates():
    try:
        duplicates = (
            db.session.query(
                StorageSystem.table_name,
                StorageSystem.storage_name,
                StorageSystem.ip_address,
                func.min(StorageSystem.id).label('min_id'),
                func.count(StorageSystem.id).label('count')
            )
            .group_by(StorageSystem.table_name, StorageSystem.storage_name, 
                     StorageSystem.ip_address)
            .having(func.count(StorageSystem.id) > 1)
            .all()
        )
        
        for dup in duplicates:
            db.session.query(StorageSystem).filter(
                StorageSystem.table_name == dup.table_name,
                StorageSystem.storage_name == dup.storage_name,
                StorageSystem.ip_address == dup.ip_address,
                StorageSystem.id != dup.min_id
            ).delete(synchronize_session=False)
        
        if duplicates:
            db.session.commit()
            logger.info(f"Удалено {len(duplicates)} дубликатов")
    except Exception as e:
        logger.error(f"Ошибка при удалении дубликатов: {e}")
        db.session.rollback()


duplicates_removed = False


@app.before_request
def remove_duplicates_once():
    global duplicates_removed
    if not duplicates_removed:
        remove_all_duplicates()
        duplicates_removed = True


# -------------------------------
# Функции Discovery для разных вендоров
# -------------------------------
def perform_discovery_by_vendor(vendor, api_username, api_password, storage_ip, storage_name, logger):
    """
    Выполняет discovery для конкретного вендора
    """
    logger.info(f"Запуск discovery для {vendor} {storage_name} ({storage_ip})")
    
    try:
        if vendor.lower() == 'huawei':
            logger.info(f"Вызов perform_huawei_discovery для {storage_name}")
            result = perform_huawei_discovery(
                api_username, api_password, storage_ip, storage_name, logger
            )
        elif vendor.lower() == 'netapp':
            logger.info(f"Вызов perform_netapp_discovery для {storage_name}")
            result = perform_netapp_discovery(
                api_username, api_password, storage_ip, storage_name, logger
            )
        elif vendor.lower() == 'ibm':
            logger.info(f"Вызов perform_ibm_discovery для {storage_name}")
            result = perform_ibm_discovery(
                api_username, api_password, storage_ip, storage_name, logger
            )
        else:
            logger.warning(f"Неизвестный вендор {vendor} для {storage_name}")
            return {
                "pools": [],
                "luns": [],
                "initiators": [],
                "storage_metrics": {storage_name: {"total_effective_tb": 0.0}}
            }
        
        # Проверяем что результат корректный
        if not isinstance(result, dict):
            logger.error(f"Некорректный тип результата от {vendor} discovery: {type(result)}")
            return {
                "pools": [],
                "luns": [],
                "initiators": [],
                "storage_metrics": {storage_name: {"total_effective_tb": 0.0}}
            }
        
        # Логируем результат для отладки
        pools_count = len(result.get("pools", []))
        luns_count = len(result.get("luns", []))
        initiators_count = len(result.get("initiators", []))
        metrics = result.get("storage_metrics", {}).get(storage_name, {})
        total_tb = metrics.get("total_effective_tb", 0)
        
        logger.info(f"Discovery результат для {vendor} {storage_name}: "
                   f"пулов={pools_count}, luns={luns_count}, инициаторов={initiators_count}, "
                   f"общая емкость={total_tb} TiB")
        
        return result
        
    except Exception as e:
        logger.exception(f"Discovery не выполнен для {vendor} {storage_name}: {str(e)}")
        return {
            "pools": [],
            "luns": [],
            "initiators": [],
            "storage_metrics": {storage_name: {"total_effective_tb": 0.0}}
        }
    

def perform_discovery_multivender(credentials_by_vendor, storages=None):
    """
    Выполняет discovery для различных типов систем хранения с разными учетными данными
    
    Args:
        credentials_by_vendor: dict с учетными данными для каждого вендора
        storages: список систем хранения
    """
    if storages is None:
        storages = StorageSystem.query.all()

    logger.info(f"Запуск Multi-vendor Discovery для {len(storages)} систем хранения")
    all_pools = []
    all_luns = []
    all_initiators = []
    storage_metrics = {}
    
    successful_discoveries = 0
    failed_discoveries = 0

    for storage in storages:
        storage_name = storage.storage_name
        storage_ip = storage.ip_address
        vendor = storage.vendor or 'Huawei'
        
        logger.info(f"Попытка подключения к {vendor} {storage_name} ({storage_ip})")
        
        # Получаем учетные данные для конкретного вендора
        vendor_credentials = credentials_by_vendor.get(vendor.lower(), {})
        api_username = vendor_credentials.get('username')
        api_password = vendor_credentials.get('password')
        
        if not api_username or not api_password:
            logger.warning(f"Отсутствуют учетные данные для {vendor} {storage_name}")
            storage_metrics[storage_name] = {"total_effective_tb": 0.0}
            failed_discoveries += 1
            continue
        
        try:
            logger.info(f"Вызов discovery для {vendor} {storage_name}")
            result = perform_discovery_by_vendor(
                vendor, api_username, api_password, storage_ip, storage_name, logger
            )
            
            if result:
                # Объединяем результаты
                result_pools = result.get("pools", [])
                result_luns = result.get("luns", [])
                result_initiators = result.get("initiators", [])
                result_metrics = result.get("storage_metrics", {})
                
                logger.info(f"Discovery результат для {storage_name}: "
                           f"pools={len(result_pools)}, luns={len(result_luns)}, "
                           f"initiators={len(result_initiators)}")
                
                all_pools.extend(result_pools)
                all_luns.extend(result_luns)
                all_initiators.extend(result_initiators)
                storage_metrics.update(result_metrics)
                
                successful_discoveries += 1
            else:
                logger.warning(f"Пустой результат discovery для {storage_name}")
                storage_metrics[storage_name] = {"total_effective_tb": 0.0}
                failed_discoveries += 1

        except Exception as e:
            logger.exception(f"Discovery не выполнен для {vendor} {storage_name} ({storage_ip}): {str(e)}")
            storage_metrics[storage_name] = {"total_effective_tb": 0.0}
            failed_discoveries += 1
            continue

    logger.info(f"Multi-vendor Discovery завершён: успешно={successful_discoveries}, "
               f"ошибок={failed_discoveries}, пулов={len(all_pools)}, "
               f"LUN/томов={len(all_luns)}, инициаторов/хостов={len(all_initiators)}")
    
    final_result = {
        "pools": all_pools,
        "luns": all_luns,
        "initiators": all_initiators,
        "storage_metrics": storage_metrics
    }
    
    # Логируем финальный результат для отладки
    logger.info(f"Финальный результат discovery: {json.dumps(final_result, indent=2, default=datetime_serializer)[:1000]}...")
    
    return final_result

# Маршрут для очистки данных discovery (для отладки)
@app.route('/clear_discovery/<table_name>')
def clear_discovery(table_name):
    """Очищает данные discovery для указанной площадки"""
    try:
        # Удаляем все записи discovery для площадки
        deleted_count = DiscoveryResult.query.filter_by(table_name=table_name).delete()
        
        # Сбрасываем время последнего обновления для систем хранения
        storages = StorageSystem.query.filter_by(table_name=table_name).all()
        for storage in storages:
            storage.last_updated = None
        
        db.session.commit()
        
        logger.info(f"Удалено {deleted_count} записей discovery для {table_name}")
        
        return redirect(url_for('discovery_form', table_filter=table_name))
        
    except Exception as e:
        db.session.rollback()
        logger.exception(f"Ошибка при очистке discovery данных: {e}")
        return f"Ошибка: {str(e)}"

# -------------------------------
# Главная страница ("/")
# -------------------------------
@app.route('/', methods=['GET', 'POST'])
def storage_form():
    message = None
    storages_list = []

    if request.method == 'POST':
        if "submit_manual" in request.form:
            table_name = request.form.get("table_name_manual")
            storage_name = request.form.get("storage_name_manual")
            ip_address = request.form.get("ip_address_manual")
            storage_role = request.form.get("storage_role_manual")
            storage_status = request.form.get("storage_status_manual")
            vendor = request.form.get("vendor_manual", "Huawei")
            
            if table_name and storage_name and ip_address:
                try:
                    existing = StorageSystem.query.filter_by(
                        table_name=table_name,
                        storage_name=storage_name,
                        ip_address=ip_address
                    ).first()
                    if not existing:
                        new_storage = StorageSystem(
                            table_name=table_name,
                            storage_name=storage_name,
                            ip_address=ip_address,
                            storage_role=storage_role,
                            storage_status=storage_status,
                            vendor=vendor,
                            last_updated=None
                        )
                        db.session.add(new_storage)
                        db.session.commit()
                        message = "Данные успешно добавлены."
                    else:
                        message = "Такая запись уже существует."
                except Exception as e:
                    db.session.rollback()
                    message = f"Ошибка при добавлении данных: {str(e)}"
            else:
                message = "Заполните все обязательные поля."

    # Получаем список площадок
    try:
        distinct_tables = [
            row[0] for row in db.session.query(StorageSystem.table_name)
            .distinct().order_by(StorageSystem.table_name).all()
        ]
    except Exception as e:
        logger.error(f"Ошибка при получении списка площадок: {e}")
        distinct_tables = []
    
    table_filter = request.args.get('table_filter', None)
    
    if table_filter is None and distinct_tables:
        table_filter = distinct_tables[0]
    
    try:
        storages_list = fetch_storages(table_filter)
    except Exception as e:
        logger.error(f"Ошибка при получении систем хранения: {e}")
        storages_list = []

    return render_template(
        "storage_form.html",
        message=message,
        storages=storages_list,
        distinct_tables=distinct_tables,
        table_filter=table_filter
    )


# -------------------------------
# Вспомогательная функция для извлечения систем хранения
# -------------------------------
def fetch_storages(table_filter):
    fetched_storages = []
    if table_filter:
        try:
            fetched_storages = (
                StorageSystem.query.filter_by(table_name=table_filter)
                .order_by(StorageSystem.vendor, StorageSystem.storage_name).all()
            )
            logger.info(
                f"Извлечено {len(fetched_storages)} систем хранения "
                f"для площадки {table_filter}"
            )
        except Exception as e:
            logger.error(f"Ошибка при извлечении систем хранения: {e}")
            fetched_storages = []
    else:
        logger.info("table_filter пустой, возвращен пустой список")
    return fetched_storages


# -------------------------------
# Синхронизация с NetBox
# -------------------------------
@app.route('/sync_netbox', methods=['GET', 'POST'])
def sync_netbox():
    if request.method == 'GET':
        table_filter = request.args.get('table_filter', '')
        return redirect(url_for('storage_form', table_filter=table_filter))

    site_param = request.form.get('site')
    api_token = request.form.get('api_token')
    message = None

    if not site_param or not api_token:
        message = "Ошибка: Укажите имя площадки и API-токен."
    else:
        try:
            netbox_storages = fetch_storage_from_netbox(site_param, api_token, NETBOX_URL)
            if not netbox_storages:
                message = "Системы хранения не найдены в NetBox."
            else:
                added_count = 0
                for storage_item in netbox_storages:
                    if ('storage_name' not in storage_item or 
                        'ip_address' not in storage_item):
                        logger.warning(
                            f"Пропущена запись из NetBox из-за отсутствия "
                            f"обязательных полей: {storage_item}"
                        )
                        continue

                    existing = StorageSystem.query.filter_by(
                        table_name=site_param,
                        storage_name=storage_item['storage_name'],
                        ip_address=storage_item['ip_address']
                    ).first()
                    if not existing:
                        new_storage = StorageSystem(
                            table_name=site_param,
                            storage_name=storage_item['storage_name'],
                            ip_address=storage_item['ip_address'],
                            storage_role=storage_item.get('storage_role'),
                            storage_status=storage_item.get('storage_status'),
                            vendor=storage_item.get('vendor', 'Huawei'),
                            last_updated=None
                        )
                        db.session.add(new_storage)
                        added_count += 1
                db.session.commit()
                message = (
                    f"Успешно добавлено {added_count} систем хранения "
                    f"из NetBox для {site_param}."
                )
                if added_count == 0 and netbox_storages:
                    message = (
                        f"Все системы хранения для {site_param} "
                        f"уже существуют в базе данных."
                    )
        except Exception as e:
            db.session.rollback()
            message = f"Ошибка при синхронизации с NetBox: {str(e)}"
            logger.exception("Ошибка синхронизации Netbox")

    table_filter_param = request.args.get('table_filter', site_param if site_param else None)
    storages_list = fetch_storages(table_filter_param)
    distinct_tables = [
        row[0] for row in db.session.query(StorageSystem.table_name).distinct().all()
    ]

    return render_template(
        "storage_form.html",
        message=message,
        storages=storages_list,
        distinct_tables=distinct_tables,
        table_filter=table_filter_param
    )


# -------------------------------
# Удаление
# -------------------------------
@app.route('/delete/<int:record_id>', methods=['POST'])
def delete_storage(record_id):
    record = StorageSystem.query.get(record_id)
    table_filter_val = None
    if record:
        table_filter_val = record.table_name
        try:
            db.session.delete(record)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return f"Ошибка при удалении записи: {str(e)}", 500
    return redirect(url_for('storage_form', table_filter=table_filter_val or ''))


# -------------------------------
# Переименование
# -------------------------------
@app.route('/rename/<int:record_id>', methods=['POST'])
def rename_storage(record_id):
    record = StorageSystem.query.get(record_id)
    table_filter_val = None
    if record:
        table_filter_val = record.table_name
        new_name = request.form.get("new_name")
        if new_name and new_name.strip():
            record.storage_name = new_name.strip()
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                return f"Ошибка при переименовании: {str(e)}", 500
    return redirect(url_for('storage_form', table_filter=table_filter_val or ''))


# -------------------------------
# Обновление информации о системе хранения
# -------------------------------
@app.route('/update_storage/<table_name>/<storage_name>', methods=['POST'])
def update_storage(table_name, storage_name):
    api_username = request.form.get('username')
    api_password = request.form.get('password')
    
    if not api_username or not api_password:
        return redirect(url_for('discovery_form', table_filter=table_name))
    
    # Находим конкретную систему хранения
    storage = StorageSystem.query.filter_by(
        table_name=table_name,
        storage_name=storage_name
    ).first()
    
    if not storage:
        return redirect(url_for('discovery_form', table_filter=table_name))
    
    try:
        # Выполняем discovery для этой конкретной системы
        vendor = storage.vendor or 'Huawei'
        result = perform_discovery_by_vendor(
            vendor, api_username, api_password, storage.ip_address, storage_name, logger
        )
        
        # Сохраняем результат в базу
        discovery_record = DiscoveryResult(
            table_name=table_name,
            data_json=json.dumps(result, default=datetime_serializer),
            timestamp=get_moscow_time()
        )
        db.session.add(discovery_record)
        
        # Обновляем время последнего обновления
        storage.last_updated = get_moscow_time()
        db.session.commit()
        
        logger.info(f"Discovery выполнен для {vendor} {storage_name}")
        
    except Exception as e:
        db.session.rollback()
        logger.exception(f"Ошибка при обновлении {storage_name}: {str(e)}")
    
    return redirect(url_for('discovery_form', table_filter=table_name))


# -------------------------------
# Форма Discovery
# -------------------------------
# Улучшенная функция discovery_form с подробной диагностикой
@app.route('/discovery', methods=['GET', 'POST'])
def discovery_form():
    table_filter = request.args.get('table_filter', '')
    iqn_search = request.args.get('search_query', '')
    wwn_search = request.args.get('search_query', '')
    search_all = request.args.get('search_all') == 'on'
    
    # Сохраняем последние введенные данные для каждого вендора
    last_credentials = {
        'huawei': {'username': '', 'password': ''},
        'netapp': {'username': '', 'password': ''},
        'ibm': {'username': '', 'password': ''}
    }
    error_message = None
    sorted_storage = []

    if request.method == 'POST':
        action = request.form.get('action')
        logger.info(f"POST запрос с action: {action}")
        
        if action == 'do_discovery':
            table_name = request.form.get('table_filter')
            logger.info(f"Discovery для площадки: {table_name}")
            
            # Собираем учетные данные для всех вендоров
            credentials_by_vendor = {
                'huawei': {
                    'username': request.form.get('huawei_username', ''),
                    'password': request.form.get('huawei_password', '')
                },
                'netapp': {
                    'username': request.form.get('netapp_username', ''),
                    'password': request.form.get('netapp_password', '')
                },
                'ibm': {
                    'username': request.form.get('ibm_username', ''),
                    'password': request.form.get('ibm_password', '')
                }
            }
            
            logger.info(f"Получены учетные данные: {[(k, bool(v['username']) and bool(v['password'])) for k, v in credentials_by_vendor.items()]}")
            
            # Сохраняем для повторного использования
            last_credentials = credentials_by_vendor
            
            # Проверяем, есть ли хотя бы одни учетные данные
            has_credentials = any(
                creds['username'] and creds['password'] 
                for creds in credentials_by_vendor.values()
            )
            
            if not has_credentials:
                error_message = "Укажите логин и пароль хотя бы для одного типа систем хранения."
                logger.warning("Нет учетных данных для discovery")
            else:
                try:
                    # Получаем системы хранения для площадки
                    storages = StorageSystem.query.filter_by(table_name=table_name).all()
                    logger.info(f"Найдено {len(storages)} систем хранения для площадки {table_name}")
                    
                    if not storages:
                        error_message = f"Системы хранения для площадки {table_name} не найдены."
                        logger.warning(f"Нет систем хранения для площадки {table_name}")
                    else:
                        # Логируем какие системы найдены
                        for storage in storages:
                            logger.info(f"Система: {storage.vendor} {storage.storage_name} ({storage.ip_address})")
                        
                        # Выполняем discovery
                        logger.info("Запуск multi-vendor discovery")
                        result = perform_discovery_multivender(credentials_by_vendor, storages)
                        
                        # Логируем результат
                        logger.info(f"Discovery результат: pools={len(result.get('pools', []))}, "
                                   f"luns={len(result.get('luns', []))}, "
                                   f"initiators={len(result.get('initiators', []))}, "
                                   f"metrics={result.get('storage_metrics', {})}")
                        
                        # Сохраняем результат
                        discovery_record = DiscoveryResult(
                            table_name=table_name,
                            data_json=json.dumps(result, default=datetime_serializer),
                            timestamp=get_moscow_time()
                        )
                        db.session.add(discovery_record)
                        
                        # Обновляем время последнего обновления для всех систем
                        for storage in storages:
                            storage.last_updated = get_moscow_time()
                        
                        db.session.commit()
                        logger.info(f"Multi-vendor Discovery сохранен для площадки {table_name}")
                        
                except Exception as e:
                    db.session.rollback()
                    error_message = f"Ошибка при выполнении Discovery: {str(e)}"
                    logger.exception(f"Ошибка Discovery для {table_name}")

    # Получаем данные для отображения
    if table_filter:
        logger.info(f"Получение данных для отображения, table_filter: {table_filter}")
        
        if iqn_search or wwn_search:
            logger.info(f"Выполняется поиск по IQN/WWN: {iqn_search or wwn_search}")
            sorted_storage = search_iqn_wwn(iqn_search, wwn_search, table_filter, search_all)
        else:
            logger.info("Получение всех данных discovery")
            sorted_storage = fetch_discovery_data(table_filter)
        
        logger.info(f"Получено {len(sorted_storage)} систем для отображения")
    else:
        logger.info("table_filter пустой, данные не получаются")

    return render_template(
        "discovery_form.html",
        sorted_storage=sorted_storage,
        table_filter=table_filter,
        iqn_search=iqn_search,
        wwn_search=wwn_search,
        error_message=error_message,
        last_huawei_username=last_credentials['huawei']['username'],
        last_huawei_password=last_credentials['huawei']['password'],
        last_netapp_username=last_credentials['netapp']['username'],
        last_netapp_password=last_credentials['netapp']['password'],
        last_ibm_username=last_credentials['ibm']['username'],
        last_ibm_password=last_credentials['ibm']['password']
    )


# Диагностическая функция для проверки данных в базе
@app.route('/debug_discovery/<table_name>')
def debug_discovery(table_name):
    """Диагностическая страница для отладки discovery данных"""
    try:
        # Проверяем системы хранения
        storages = StorageSystem.query.filter_by(table_name=table_name).all()
        storages_info = []
        for storage in storages:
            storages_info.append({
                'name': storage.storage_name,
                'vendor': storage.vendor,
                'ip': storage.ip_address,
                'last_updated': storage.last_updated
            })
        
        # Проверяем записи discovery
        discovery_records = (
            DiscoveryResult.query
            .filter_by(table_name=table_name)
            .order_by(DiscoveryResult.timestamp.desc())
            .limit(5)
            .all()
        )
        
        discovery_info = []
        for record in discovery_records:
            try:
                data = json.loads(record.data_json)
                discovery_info.append({
                    'timestamp': record.timestamp,
                    'pools_count': len(data.get('pools', [])),
                    'luns_count': len(data.get('luns', [])),
                    'initiators_count': len(data.get('initiators', [])),
                    'storage_metrics': data.get('storage_metrics', {})
                })
            except Exception as e:
                discovery_info.append({
                    'timestamp': record.timestamp,
                    'error': str(e)
                })
        
        debug_info = {
            'table_name': table_name,
            'storages_count': len(storages),
            'storages': storages_info,
            'discovery_records_count': len(discovery_records),
            'discovery_records': discovery_info
        }
        
        return f"<h1>Debug info for {table_name}</h1><pre>{json.dumps(debug_info, indent=2, default=str)}</pre>"
        
    except Exception as e:
        return f"<h1>Debug error</h1><pre>{str(e)}</pre>"


# Упрощенная тестовая функция discovery
@app.route('/test_discovery/<table_name>')
def test_discovery(table_name):
    """Тестовая функция для проверки discovery без учетных данных"""
    try:
        logger.info(f"Тест discovery для {table_name}")
        
        # Создаем тестовые данные
        test_result = {
            "pools": [
                {
                    "storage_name": "test-storage",
                    "pool_name": "test-pool",
                    "dataspace": 100.0,
                    "used_capacity": 50.0,
                    "free_capacity": 50.0,
                    "subscribed_capacity": 75.0
                }
            ],
            "luns": [
                {
                    "storage_name": "test-storage",
                    "name": "test-lun",
                    "size": 10.0,
                    "wwn": "test-wwn",
                    "pool_id": "test-pool-id"
                }
            ],
            "initiators": [
                {
                    "storage_name": "test-storage",
                    "iqn": "test-iqn",
                    "host_name": "test-host",
                    "init_status": "Online",
                    "hostIP": "192.168.1.100"
                }
            ],
            "storage_metrics": {
                "test-storage": {
                    "total_effective_tb": 100.0
                }
            }
        }
        
        # Сохраняем тестовый результат
        discovery_record = DiscoveryResult(
            table_name=table_name,
            data_json=json.dumps(test_result, default=datetime_serializer),
            timestamp=get_moscow_time()
        )
        db.session.add(discovery_record)
        db.session.commit()
        
        logger.info(f"Тестовые данные сохранены для {table_name}")
        
        return redirect(url_for('discovery_form', table_filter=table_name))
        
    except Exception as e:
        logger.exception(f"Ошибка в test_discovery: {e}")
        return f"Ошибка: {str(e)}"

# -------------------------------
# Поиск по IQN/WWN
# -------------------------------
def search_iqn_wwn(search_query, wwn_query, table_filter, search_all=False):
    """Поиск по IQN инициаторов или WWN LUN"""
    if not search_query:
        return []

    # Определяем в каких таблицах искать
    tables_to_search = []
    if search_all:
        distinct_tables = [
            row[0] for row in db.session.query(StorageSystem.table_name).distinct().all()
        ]
        tables_to_search = distinct_tables
    else:
        tables_to_search = [table_filter] if table_filter else []

    found_storage = []
    
    for table_name in tables_to_search:
        latest_record = (
            DiscoveryResult.query
            .filter_by(table_name=table_name)
            .order_by(DiscoveryResult.timestamp.desc())
            .first()
        )
        
        if not latest_record:
            continue
            
        try:
            data = json.loads(latest_record.data_json)
            
            # Поиск среди инициаторов по IQN
            for initiator in data.get('initiators', []):
                if search_query.lower() in initiator.get('iqn', '').lower():
                    storage_name = initiator.get('storage_name', '')
                    if storage_name not in [s[0] for s in found_storage]:
                        storage_data = build_storage_data(storage_name, data, table_name)
                        found_storage.append((storage_name, storage_data))
            
            # Поиск среди LUN по WWN
            for lun in data.get('luns', []):
                if search_query.lower() in lun.get('wwn', '').lower():
                    storage_name = lun.get('storage_name', '')
                    if storage_name not in [s[0] for s in found_storage]:
                        storage_data = build_storage_data(storage_name, data, table_name)
                        found_storage.append((storage_name, storage_data))
        
        except Exception as e:
            logger.exception(f"Ошибка при поиске в данных для {table_name}: {str(e)}")
            continue
    
    return found_storage


# -------------------------------
# Получение данных Discovery
# -------------------------------
# Улучшенная функция fetch_discovery_data в app.py
def fetch_discovery_data(table_filter):
    """Получает последние данные discovery для указанной площадки"""
    try:
        logger.info(f"Получение данных discovery для площадки {table_filter}")
        
        latest_record = (
            DiscoveryResult.query
            .filter_by(table_name=table_filter)
            .order_by(DiscoveryResult.timestamp.desc())
            .first()
        )
        
        if not latest_record:
            logger.warning(f"Нет записей discovery для площадки {table_filter}")
            return []
        
        logger.info(f"Найдена запись discovery от {latest_record.timestamp}")
        
        data = json.loads(latest_record.data_json)
        logger.info(f"Данные загружены: pools={len(data.get('pools', []))}, "
                   f"luns={len(data.get('luns', []))}, "
                   f"initiators={len(data.get('initiators', []))}")
        
        # Логируем метрики для отладки
        storage_metrics = data.get('storage_metrics', {})
        logger.info(f"Storage metrics: {storage_metrics}")
        
        storage_names = set()
        
        # Собираем уникальные имена систем хранения
        for item in (data.get('pools', []) + data.get('luns', []) + data.get('initiators', [])):
            storage_name = item.get('storage_name')
            if storage_name:
                storage_names.add(storage_name)
        
        logger.info(f"Найдено систем хранения: {list(storage_names)}")
        
        sorted_storage = []
        for storage_name in sorted(storage_names):
            logger.info(f"Обработка данных для {storage_name}")
            storage_data = build_storage_data(storage_name, data, table_filter)
            sorted_storage.append((storage_name, storage_data))
        
        logger.info(f"Обработано {len(sorted_storage)} систем хранения")
        return sorted_storage
        
    except Exception as e:
        logger.exception(f"Ошибка при обработке данных discovery: {e}")
        return []


# -------------------------------
# Построение данных о системе хранения
# -------------------------------
def build_storage_data(storage_name, data, table_filter):
    """Строит структуру данных для отображения системы хранения"""
    try:
        # Получаем информацию о системе из базы
        storage_info = StorageSystem.query.filter_by(
            table_name=table_filter,
            storage_name=storage_name
        ).first()
        
        # Фильтруем данные по имени системы хранения
        pools = [p for p in data.get('pools', []) if p.get('storage_name') == storage_name]
        luns = [l for l in data.get('luns', []) if l.get('storage_name') == storage_name]
        # # Повторная сортировка LUN/томов по агрегату
        # luns.sort(key=lambda x: (x.get('aggregate') or x.get('pool') or x.get('pool_id') or ''))
        initiators = [i for i in data.get('initiators', []) if i.get('storage_name') == storage_name]
        
    # ---------- новая часть: карта «агрегат → 0 / 1» ----------
        aggregate_color_map = {}
        switch = 0
        for lun in luns:                         # luns уже отсортированы по aggregate
            agg = lun.get('aggregate')
            if agg not in aggregate_color_map:
                aggregate_color_map[agg] = switch % 2
                switch += 1

        # Логируем для отладки
        logger.info(f"build_storage_data для {storage_name}: "
                   f"pools={len(pools)}, luns={len(luns)}, initiators={len(initiators)}")
        
        # Рассчитываем метрики
        storage_metrics = data.get('storage_metrics', {}).get(storage_name, {})
        total_effective_tb = storage_metrics.get('total_effective_tb', 0)
        
        logger.info(f"Метрики для {storage_name}: total_effective_tb={total_effective_tb}")
        
        # Рассчитываем суммарную подписку
        sum_subscribed_capacity_tb = 0
        for pool in pools:
            subscribed = pool.get('subscribed_capacity', 0)
            try:
                sum_subscribed_capacity_tb += float(subscribed)
            except (ValueError, TypeError):
                logger.warning(f"Некорректное значение subscribed_capacity для пула {pool.get('pool_name', '')}: {subscribed}")
        
        subscription_ratio = 'N/A'
        if total_effective_tb > 0:
            subscription_ratio = f"{(sum_subscribed_capacity_tb / total_effective_tb * 100):.2f}"
        
        logger.info(f"Подписка для {storage_name}: {sum_subscribed_capacity_tb}/{total_effective_tb} = {subscription_ratio}%")
        
        # Определяем цвет на основе подписки
        color = 'grey'
        if subscription_ratio != 'N/A':
            try:
                ratio_float = float(subscription_ratio)
                if ratio_float < 80:
                    color = 'green'
                elif ratio_float < 95:
                    color = 'orange'
                else:
                    color = 'red'
            except ValueError:
                color = 'grey'
        
        return {
            'ip_address': storage_info.ip_address if storage_info else '',
            'storage_role': storage_info.storage_role if storage_info else '',
            'storage_status': storage_info.storage_status if storage_info else '',
            'vendor': storage_info.vendor if storage_info else 'Unknown',
            'last_updated': storage_info.last_updated if storage_info else None,
            'pools': pools,
            'luns': luns,
            'initiators': initiators,
            'total_effective_tb': total_effective_tb,
            'sum_subscribed_capacity_tb': sum_subscribed_capacity_tb,
            'subscription_ratio': subscription_ratio,
            'color': color,
            'aggregate_color_map': aggregate_color_map   # ← передаём в шаблон
        }
    except Exception as e:
        logger.error(f"Ошибка при построении данных для {storage_name}: {e}")
        return {
            'ip_address': '',
            'storage_role': '',
            'storage_status': '',
            'vendor': 'Unknown',
            'last_updated': None,
            'pools': [],
            'luns': [],
            'initiators': [],
            'total_effective_tb': 0,
            'sum_subscribed_capacity_tb': 0,
            'subscription_ratio': 'N/A',
            'color': 'grey'
        }

# -------------------------------
# Скачивание Excel
# -------------------------------
@app.route('/download_excel', methods=['POST'])
def download_excel():
    table_filter = request.form.get('table_filter')
    if not table_filter:
        return redirect(url_for('storage_form'))
    
    try:
        # Получаем данные
        sorted_storage = fetch_discovery_data(table_filter)
        
        if not sorted_storage:
            return render_template("export_result.html", 
                                 message="Нет данных для экспорта.")
        
        # Создаем Excel файл
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        
        # Лист "Пулы"
        ws_pools = wb.active
        ws_pools.title = "Pools"
        pools_headers = ["Storage Name", "Vendor", "Pool Name", "Dataspace (TiB)", "Used Capacity (TiB)", "Free Capacity (TiB)", "Subscribed Capacity (TiB)"]
        ws_pools.append(pools_headers)
        
        for row in ws_pools[1:1]:
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for storage_name, data in sorted_storage:
            for pool in data['pools']:
                ws_pools.append([
                    storage_name,
                    data.get('vendor', 'Unknown'),
                    pool.get('pool_name', ''),
                    pool.get('dataspace', 0),
                    pool.get('used_capacity', 0),
                    pool.get('free_capacity', 0),
                    pool.get('subscribed_capacity', 0)
                ])
        
        # Лист "LUN/Volumes"
        ws_luns = wb.create_sheet("LUNs_Volumes")
        luns_headers = ["Storage Name", "Vendor", "LUN/Volume Name", "WWN/UUID", "Size (TiB)", "Pool/Aggregate", "SVM"]
        ws_luns.append(luns_headers)
        
        for row in ws_luns[1:1]:
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for storage_name, data in sorted_storage:
            for lun in data['luns']:
                ws_luns.append([
                    storage_name,
                    data.get('vendor', 'Unknown'),
                    lun.get('name', ''),
                    lun.get('wwn', ''),
                    lun.get('size', 0),
                    lun.get('aggregate', lun.get('pool', lun.get('pool_id', ''))),
                    lun.get('svm', '')
                ])
        
        # Лист "Инициаторы/Хосты/SVM"
        ws_initiators = wb.create_sheet("Initiators_Hosts_SVM")
        initiators_headers = ["Storage Name", "Vendor", "IQN/Name", "Host Name", "Status", "Host IP", "Type", "UUID"]
        ws_initiators.append(initiators_headers)
        
        for row in ws_initiators[1:1]:
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for storage_name, data in sorted_storage:
            for initiator in data['initiators']:
                ws_initiators.append([
                    storage_name,
                    data.get('vendor', 'Unknown'),
                    initiator.get('iqn', initiator.get('name', '')),
                    initiator.get('host_name', ''),
                    initiator.get('init_status', initiator.get('state', '')),
                    initiator.get('hostIP', ''),
                    initiator.get('type', ''),
                    initiator.get('uuid', '')
                ])
        
        # Сохраняем файл
        filename = f"storage_discovery_{table_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)
        
        return render_template("export_result.html", 
                             message=f"Файл {filename} успешно сохранён в {filepath}")
        
    except Exception as e:
        logger.exception(f"Ошибка при создании Excel файла: {str(e)}")
        return render_template("export_result.html", 
                             message=f"Ошибка при создании файла: {str(e)}")


# -------------------------------
# Блокировка двойного запуска вкладки
# -------------------------------
def open_browser_once(url):
    if os.path.exists(LOCKFILE):
        print("Lockfile существует, браузер уже был открыт или открывается.")
        return
    try:
        with open(LOCKFILE, 'w') as f:
            f.write('1')
        webbrowser.open(url)
        t = threading.Thread(target=_remove_lock_after_delay, daemon=True)
        t.start()
    except Exception as e:
        print(f"Не удалось создать lock-файл или открыть браузер: {e}")


def _remove_lock_file():
    if os.path.exists(LOCKFILE):
        try:
            os.remove(LOCKFILE)
            print("Lock-файл удален при выходе.")
        except Exception as e:
             print(f"Ошибка при удалении lock-файла: {e}")


def _remove_lock_after_delay(delay=10):
    time.sleep(delay)
    if os.path.exists(LOCKFILE):
        try:
            os.remove(LOCKFILE)
            print(f"Lock-файл удален по таймеру ({delay} секунд).")
        except Exception as e:
            print(f"Ошибка при удалении lock-файла по таймеру: {e}")


# -------------------------------
# Запуск приложения
# -------------------------------
if __name__ == '__main__':
    # Удаляем lock-файл при выходе
    atexit.register(_remove_lock_file)
    
    # Открываем браузер только один раз
    threading.Timer(1.0, lambda: open_browser_once("http://127.0.0.1:5000")).start()
    
    try:
        app.run(debug=False, host='127.0.0.1', port=5000, use_reloader=False)
    except KeyboardInterrupt:
        print("\nПриложение остановлено пользователем.")
    finally:
        _remove_lock_file()